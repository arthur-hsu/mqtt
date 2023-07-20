import os, time, sys, datetime, re
import openpyxl
from openpyxl.styles import Alignment , Font





class python_excel():
    def __init__(self, excel_path, avg_list, uplink_intv):
        avg_list = [item.lower() for item in avg_list]
        self.excel_path = excel_path
        self.uplink_intv = uplink_intv
        self.workbook = openpyxl.load_workbook(excel_path)
        self.result_sheet = self.workbook['RESULT']
        self.result_sheet.cell(row=1,column= self.result_sheet.max_column+1).value='Note'
        self.result_title = [cell.value.lower() for cell in self.result_sheet[1]]# [1] is row               # title_name
        self.result_title_index = {item : self.result_title.index(item)+1 for item in self.result_title}    # title_index
        self.dataloss, self.Miss_FCnt_list , self.FCnt_recurrence=0, [], []

        self.remain_columns = [col for col in self.result_title if col in avg_list]                         # excel 裏存在的columns
        self.remain_columns_dict ={'%s_data_list'%col:[] for col in self.remain_columns}                    # 將存在的標題 按照標題將data存成list
    def read(self,):
        for row in range(3,self.result_sheet.max_row+1):
            TimeMsg, FCntMsg= '',''
            #### FCnt ######################################
            current_fcnt = self.result_sheet.cell(row, self.result_title_index['fcnt']).value
            last_fcnt = self.result_sheet.cell(row-1, self.result_title_index['fcnt']).value
            if current_fcnt == last_fcnt:
                self.FCnt_recurrence.append(str(current_fcnt))
                FCntMsg = 'FCnt %s recurrence.'% current_fcnt
            elif current_fcnt-1 != last_fcnt:
                if current_fcnt-1 == last_fcnt+1 : Miss_FCnt = int(current_fcnt-1)
                else : Miss_FCnt ='%s~%s'%(int(last_fcnt+1), int(current_fcnt-1))
                FCntMsg ='Miss FCnt %s packet.'%Miss_FCnt
                self.dataloss += current_fcnt - last_fcnt
                self.Miss_FCnt_list.append(str(Miss_FCnt))
            
            #### Time ######################################
            current_time = datetime.datetime.strptime(self.result_sheet.cell(row, self.result_title_index['time']).value, "%Y-%m-%d %H:%M:%S")
            last_time = datetime.datetime.strptime(self.result_sheet.cell(row-1, self.result_title_index['time']).value, "%Y-%m-%d %H:%M:%S")
            time_diff = abs(current_time - last_time)
            if time_diff> datetime.timedelta(seconds=self.uplink_intv*1.5):
                # print('  %s %s\n- %s %s\n--------------------------\n=                  %s\n'%(current_fcnt, current_time, last_fcnt, last_time, time_diff))
                TimeMsg = 'Haven\'t received data in %s.'%time_diff
            ################################################

            for col in self.remain_columns:
                if self.result_sheet.cell(row, self.result_title_index[col]).value is not None:
                    self.remain_columns_dict['%s_data_list'%col].append(float(self.result_sheet.cell(row, self.result_title_index[col]).value)) 
            if (TimeMsg or FCntMsg)!='':
                if (TimeMsg and FCntMsg) !='': note_str = '1. %s\n2. %s'%(FCntMsg, TimeMsg)
                elif FCntMsg != '' : note_str = FCntMsg
                elif TimeMsg != '' : note_str = TimeMsg
                self.result_sheet.cell(row=row,column= self.result_title_index['note']).value = note_str
        # print(self.remain_columns_dict)
        

    def create_report_sheet(self):
        
        self.workbook.create_sheet('REPORT',index = 0)
        self.report_sheet = self.workbook['REPORT']
        self.report_sheet['A1'] = 'REPORT'
        fcnt_begin = self.result_sheet.cell(row=2,column=self.result_title_index['fcnt']).value
        fcnt_end = self.result_sheet.cell(row=self.result_sheet.max_row , column=self.result_title_index['fcnt']).value
        fcnt_list = ['FCnt', 
                     'Total: %s'%(fcnt_end-fcnt_begin),
                     'Dataloss: %s'%self.dataloss,
                     'Success ratio: {:.2%}'.format((fcnt_end-self.dataloss)/fcnt_end),
                     'Begin: %s'%fcnt_begin,
                     'End: %s'%fcnt_end,
                     'Miss FCnt number\n%s'%self.Miss_FCnt_list,
                     'FCnt recurrence\n%s'%self.FCnt_recurrence]
        for i in range(1,len(fcnt_list)+1) : self.report_sheet.cell(row=2, column=i).value = fcnt_list[i-1]
        self.report_sheet['A3'] = 'Average'
        avg_list=['{} avg: {:.1f}'.format(col, sum(self.remain_columns_dict['%s_data_list'%col])/len(self.remain_columns_dict['%s_data_list'%col])) for col in self.remain_columns]
        for i in range(2,len(avg_list)+2) : self.report_sheet.cell(row=3, column=i).value = avg_list[i-2]
        self.workbook.save(self.excel_path)
        
        print(f'{avg_list=}')
        print(f'{self.remain_columns=}')
    def set_wedth(self, excel_path):
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.sheetnames
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # 取得欄位的字母表示法
                for cell in col:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                adjusted_width = (max_length + 2) * 1.2  # 計算自動調整後的欄寬
                worksheet.column_dimensions[column].width = adjusted_width  # 設定欄寬
            for row in worksheet.rows:
                for cell in row:
                    # cell.font = font
                    cell.alignment = Alignment(horizontal='center',vertical='center')
        # RESULT ############################
        workbook['RESULT'].column_dimensions['H'].width = 35
        for i in range(1, self.result_sheet.max_row):
            workbook['RESULT'].cell(row=i,column=self.result_title_index['note']).alignment = Alignment(horizontal='left',vertical='top',wrapText=True)
        for cell in workbook['RESULT'][1]: 
            cell.font= Font(bold=True)
            cell.alignment = Alignment(horizontal='center',vertical='center')
        workbook['RESULT'].freeze_panes = 'B1'
        # REPORT ############################
        for row in workbook['REPORT']:
            for cell in row: cell.font= Font(bold=True)
        workbook['REPORT'].column_dimensions['G'].width = 75
        workbook['REPORT']['G2'].alignment = Alignment(horizontal='left',vertical='top',wrapText=True)
        workbook['REPORT'].column_dimensions['H'].width = 75
        workbook['REPORT']['H2'].alignment = Alignment(horizontal='left',vertical='top',wrapText=True)
        workbook['REPORT'].merge_cells('A1:H1')
        
        workbook.save(self.excel_path)
        workbook.close()


    def test(self):
        self.read()
        self.create_report_sheet()
        self.set_wedth(self.excel_path)


if __name__ == '__main__':
    original_excel  = '/home/arthur/Desktop/mqtt_report.xlsx'
    avg_list =['rssi','snr','temperature','humid'] 
    excel = python_excel(original_excel,avg_list, 180)
    excel.test()
