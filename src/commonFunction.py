#! python3
#----- import built-in Module -----#
import re
import time
import os, platform, subprocess, queue, math
import locale
import requests
operation_system = platform.system()

#----- import 3rd-party Module -----#
import pyscreenshot as ImageGrab
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl import drawing
import pandas as pd
from prettytable import PrettyTable
from PIL import Image
import pytesseract

#----- import Self Module -----#
import config.config as cf
config_path = os.path.abspath(os.path.dirname(__file__)) + '/config/'
from . import colorLog as log

encodeType = locale.getdefaultlocale()[1]
stopScript = 0

class Common:
    def __init__(self):
        None

    def net_info(self):
        if 'Windows' in operation_system:
            cmd = 'ipconfig'
        else:
            cmd = 'ifconfig' 
        net_info = {}
        command = os.popen(cmd)
        while True:
            tmp = command.buffer.readline().decode(locale.getdefaultlocale()[1], 'ignore')
            if tmp == '':
                break
        ### mac
            elif ': flags=' in tmp :
                iface = tmp.split(': flags=')[0]
                info = ''
                breakString = ['status:', 'nd6 options=']
                for i in range(20):
                    tmp = command.buffer.readline().decode(locale.getdefaultlocale()[1], 'ignore')
                    info += tmp
                    if any(string in tmp for string in breakString) :
                        break
                net_info[iface] = info
        ### windows
            elif 'adapter ' in tmp or '網路卡 ' in tmp or '介面卡 ' in tmp:
                iface = tmp.split(' ')
                iface.pop(0)
                iface = ' '.join(iface).replace(':', '').replace('\r\n', '')
                info = ''
                breakString = 0
                for i in range(20):
                    tmp = command.buffer.readline().decode(locale.getdefaultlocale()[1], 'ignore')
                    info += tmp
                    if '\r\n' in tmp and breakString == 0 :
                        breakString = 1
                    elif '\r\n' == tmp or '' == tmp:
                        break
                net_info[iface] = info
        return net_info
    

    def parsing_img_into_text(self, file_path):
        try:
            img = Image.open(file_path)
            text = pytesseract.image_to_string(img, lang='eng')
            log.Logger('\n\n==========================  parsing image into text ==========================\n', timestamp=0)
            log.Logger(text, timestamp=0)
            log.Logger('==========================  parsing image end ==========================\n\n', timestamp=0)
        except:
            None



    def screen_shot(self, name=''):
        image = ImageGrab.grab()
        shotTime = time.strftime( "%Y-%m-%d %X", time.localtime() )
        shotTime = shotTime.replace(":", "")
        shotTime = shotTime.replace(" ", ".")
        path = cf.get_value('screenshot_path')
        saveto = path + '%s_%s.png' %(str(shotTime), name.replace(':','.').replace('?',''))
        image.save(saveto)
        #self.parsing_img_into_text(saveto)
        return saveto
    
    def shell(self, cmd, expected='', shell=True):
        if shell == True:
            log.Logger('%s' %cmd, 'BLACK', 'WHITE')
        p = os.popen(cmd)
        dump = ''
        result = False
        if expected != '' and type(expected) != type(['test']):
            expected = [expected]
        while True:
            line = p.buffer.readline().decode(encodeType, 'ignore')
            #line = p.buffer.readline().decode('ascii', 'ignore')
            if line == '':
                break
            else:
                line = line.strip(' ').replace('\n','').replace('\r','')
                if line != '' and shell==True:
                    log.Logger(line)
                if expected != '':
                    if any(string in line for string in expected):
                        result = line    
                else:
                    dump += '%s\n' %line
        p.close()
        if expected == '':
            result = dump
        return result


    def ping(self, ip):
        global stopScript
        if 'Windows' in platform.system():
            cmd = 'ping ' + ip + ' -n 5'
            ttl = 'TTL='
        else:
            cmd = 'ping ' + ip + ' -c 5'
            ttl = 'ttl='
        
        bPing = False
        for i in range(5):
            if stopScript == 1 :
                return False
            ping = os.popen(cmd)
            for i in range(5):
                tmp = ping.buffer.readline().decode(locale.getdefaultlocale()[1], 'ignore')
                log.Logger(tmp.replace('\n','').replace('\r',''))
                if ttl in tmp :
                    bPing = True
                    time.sleep(5-i)
                    return bPing
                else:
                    time.sleep(1)
        return bPing

    def Table_Insert_Data(self, row_list):
        self.table = cf.get_value('report_table')
        self.table.add_row(row_list)
        cf.set_value('report_table', self.table)
    
    def Filter_Number(self, string):
        handle = string
        handle = filter(str.isdigit, handle)
        handle = ''.join(list(handle))
        return int(handle)


    def Optimized_Excel(self, excelPath):
        self.excelPath = excelPath
        wb = load_workbook(self.excelPath)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            # 凍結第一列
            ws.freeze_panes = 'B2'
            # 調整列寬
            df = pd.read_excel(self.excelPath, sheet_name=sheet, engine='openpyxl')
            maxcolumn = df.shape[1]
            maxrow = df.shape[0]
            for col in df.columns:
                # 獲取列序號
                index = list(df.columns).index(col)
                # 獲取行字母開頭
                letter = get_column_letter(index+1)
                # 獲取當前最大寬度
                collen = df[col].apply(lambda x :len(str(x).encode())).max()
                #print(str(col), str(collen))
                # 設置寬度為最大字幕長度
                if collen < 17:
                    collen = 17
                ws.column_dimensions[letter].width = collen + 1

            # 設定邊框
            bian = Side(style='thin', color='000000')    # 設定邊框樣式
            border = Border(top=bian, bottom=bian, left=bian, right=bian)

            font = Font(size = 12)
            font_green = Font(color = '008000')
            ali_center = Alignment(horizontal='center', vertical='center') # 置中
            ali_left = Alignment(horizontal='left', vertical='center') # 靠左
            yellow = PatternFill('solid', fgColor='c1ddff') # 儲存格淺黃色
            pink = PatternFill('solid', fgColor='ffe4e1') # 儲存格淺粉色
            red = PatternFill('solid', fgColor='8b0000') # 儲存格暗紅色
            grey = PatternFill('solid', fgColor='a9a9a9') # 儲存灰色

            for col in df.columns: 
                if col == '$Description':
                    # 獲取列序號
                    index = list(df.columns).index(col)
                    for row in ws.iter_rows(min_row=2, max_row=maxrow+1, min_col=index+1, max_col=index+1): # 遍歷第1欄儲存格, 看到結果標黃底
                            for cell in row:
                                cell.alignment = ali_left
                                #data = ws.cell(cell.row, cell.column-(i+1)).value

            for row in ws.iter_rows(min_row=0, max_row=maxrow+1, min_col=0, max_col=maxcolumn): # 遍歷所有儲存格, 調整字型, 文字置中, 框線
                    for cell in row:
                            cell.font = font
                            if cell.alignment != ali_left:
                                cell.alignment = ali_center
                            cell.border = border
                            if '*' in str(cell.value) and str(cell.value).index('*') == 0:                                   ###看到含有*標紅底
                                cell.fill = red
                                try:
                                    cell.value = float(cell.value.replace('*', ''))
                                except:
                                    cell.value = cell.value.replace('*', '')
                            elif '$' in str(cell.value) and str(cell.value).index('$') == 0:                                 ###看到含有$標灰底
                                cell.fill = grey
                                cell.value = cell.value.replace('$', '')
                            elif 'Pass' in str(cell.value):
                                cell.font = font_green
            '''if ali_left_col != '':
                for row in ws.iter_rows(min_row=1, max_row=maxrow+1, min_col=ali_left_col, max_col=ali_left_col): # 遍歷第一欄儲存格, 文字靠左
                        for cell in row:
                                cell.alignment = ali_left'''
            
            
            wb.save(excelPath)

    def WaitTime(self, WaitTime):
        WaitTime = int(WaitTime)
        print('\n')
        for sec in range(0, WaitTime):
            print(f"Please wait {WaitTime} seconds... < Timer: {WaitTime - sec} >      ", "\r" , end='')
            time.sleep(1)


    def Line_Chart(self, excelPath, title, x_title, y_title, start_data_col, end_data_col, add_position):
        wb = load_workbook(excelPath)
        ws = wb[wb.sheetnames[0]]
        lineChat = LineChart()
        #lineChat.title = "5G_CH40_TX"
        lineChat.title = title
        lineChat.style = 11
        lineChat.x_axis.title = x_title
        lineChat.y_axis.title = y_title
        lineChat.y_axis.scaling.min = 0
        

        df = pd.read_excel(excelPath, sheet_name=wb.sheetnames[0], engine='openpyxl')
        maxrow = df.shape[0]
        # 參照值
        x_axis_data = Reference(ws, min_row=2, max_row=maxrow+1, min_col=1, max_col=1)
        data = Reference(ws, min_row=1, max_row=maxrow+1, min_col=start_data_col, max_col=end_data_col)
        # 將資料加入圖形
        lineChat.add_data(data, titles_from_data=True)
        # 設定 X 軸標示值
        lineChat.set_categories(x_axis_data)
        # Style the lines
        Fill = ['blue', 'green', 'red', 'grey']
        for i in range(len(lineChat.series)):
            lineChat.series[i].smooth = True
            lineChat.series[i].graphicalProperties.line =  drawing.line.LineProperties(solidFill = drawing.colors.ColorChoice(prstClr= Fill[i]))
            lineChat.series[i].graphicalProperties.line.width = 40000 # make the line thicker than normal
        # 將圖形放置在  儲存格位置
        ws.add_chart(lineChat, add_position)
        wb.save(excelPath)




    
def raise_exception():
    print('Got the raise exception, stop the script.')
    stopScript = 1



def string_to_list_remove_empty(string, handle):
#### .strip(' ') 去掉最左右兩邊的指定符號
    string = string.strip(' ').split(handle)
### remove empty from list
    string = list(filter(None, string))
    return string

    
def killProcess(keyword):
    log.Logger( 'Kill Process : ' + keyword, 'BLACK', 'WHITE', timestamp=0)
    if 'Windows' in operation_system:
    ### Find the PID with port.
        if '-p' in keyword:
            handle = string_to_list_remove_empty(keyword, ' ')
            command = os.popen('netstat -ano | findstr ' + handle[handle.index('-p') + 1])
            port_PID = command.buffer.readline().decode('ascii', 'ignore')
    ### Find the PID with name.
            command = os.popen('tasklist | find "' + handle[0] + '.exe"')
            tmp = command.buffer.readline().decode('ascii', 'ignore')
            name_PID = string_to_list_remove_empty(tmp, ' ')
            if name_PID[1] in port_PID:
                print(tmp)
                command = os.popen('taskkill /PID ' + name_PID[1] + ' /F')
            else:
                print('Can not find ' + keyword + ' and port ' + handle[handle.index('-p') + 1])
        else:
            log.Logger('Taskkill /F /IM ' + keyword + '.exe')
            command = os.popen('Taskkill /F /IM ' + keyword + '.exe')
    else:
        command = os.popen('ps aux | grep "' + keyword + '"')
        while True:
            tmp = command.buffer.readline().decode('ascii', 'ignore')
            items = string_to_list_remove_empty(tmp, ' ')
            if tmp == '':
                break
            elif (len(items) > 3) and ('mount' not in tmp) and ('grep' not in tmp) and ('Studio' not in tmp) and ('/Applications/Vysor.app/Contents/Frameworks/Vysor' not in tmp) :
                print(tmp.replace('\n', ''))
                print("kill {0}...".format(items[1], subprocess.call(["kill", '-9',items[1]])))


'''def grep_IP_Self():
    localIP = []
    for ip in socket.gethostbyname_ex(socket.gethostname())[2]:
        if '169.254' not in ip:
            localIP.append(ip)
    log.Logger('IP List: %s' %localIP)
    return localIP'''

def grep_IP_Self(IPver='IPv4'):
    if 'Windows' in operation_system:
        cmd = 'ipconfig | find "' + IPver + '"'
        position = -1
    else:
        IPver = IPver.replace('IPv4', 'inet ').replace('IPv6', 'inet6 ')
        cmd = 'ifconfig | grep "' + IPver + '"'

    command = os.popen(cmd)
    localIP = ['127.0.0.1']
    for i in range(20):
        tmp = command.buffer.readline().decode(locale.getdefaultlocale()[1], 'ignore')
        if tmp == '':
            break
        elif '%' in tmp:
            continue
        else:
            tmp = tmp.replace('\n', '').replace('\r', '').replace('\t','')
            log.Logger(tmp, timestamp=0)
            handle = tmp.split(' ')
            if 'Windows' not in operation_system:
                position = handle.index(IPver.replace(' ',''))+1
            ip = handle[position]
            filter_list = ['169.254.', '127.0.0.1']
            if any(string in tmp for string in filter_list) or ip == '':
                pass
            else:
                localIP.append(ip)
    return localIP
